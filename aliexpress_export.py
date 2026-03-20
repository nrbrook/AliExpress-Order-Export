# /// script
# requires-python = ">=3.13,<3.14"
# dependencies = [
#   "browser-cookie3>=0.20.1,<1",
#   "openpyxl>=3.1.5,<4",
#   "pillow>=11.2,<12",
# ]
# ///

from __future__ import annotations

import argparse
import base64
import configparser
import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
import subprocess
import tempfile
import time
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Optional
from urllib.parse import parse_qs, urlencode, urlparse

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DEFAULT_EXPORT_DIR = "exports"
DEFAULT_AUTH_DIR = ".auth"
DEFAULT_COOKIES_PATH = f"{DEFAULT_AUTH_DIR}/cookies.json"
DEFAULT_MAX_PAGES = 50
DEFAULT_PAGE_SIZE = 10
DEFAULT_API_APP_KEY = "12574478"
DEFAULT_ORDERS_PAGE_URL = "https://www.aliexpress.com/p/order/index.html"
SUPPORTED_BROWSERS = {
    "chrome": {
        "label": "Google Chrome",
        "app_paths": [
            "/Applications/Google Chrome.app",
            "~/Applications/Google Chrome.app",
        ],
        "bundle_id": "com.google.Chrome",
        "support_dir": "~/Library/Application Support/Google/Chrome",
        "family": "chromium",
    },
    "brave": {
        "label": "Brave",
        "app_paths": [
            "/Applications/Brave Browser.app",
            "~/Applications/Brave Browser.app",
        ],
        "bundle_id": "com.brave.Browser",
        "support_dir": "~/Library/Application Support/BraveSoftware/Brave-Browser",
        "family": "chromium",
    },
    "edge": {
        "label": "Microsoft Edge",
        "app_paths": [
            "/Applications/Microsoft Edge.app",
            "~/Applications/Microsoft Edge.app",
        ],
        "bundle_id": "com.microsoft.edgemac",
        "support_dir": "~/Library/Application Support/Microsoft Edge",
        "family": "chromium",
    },
    "chromium": {
        "label": "Chromium",
        "app_paths": [
            "/Applications/Chromium.app",
            "~/Applications/Chromium.app",
        ],
        "bundle_id": "org.chromium.Chromium",
        "support_dir": "~/Library/Application Support/Chromium",
        "family": "chromium",
    },
    "vivaldi": {
        "label": "Vivaldi",
        "app_paths": [
            "/Applications/Vivaldi.app",
            "~/Applications/Vivaldi.app",
        ],
        "bundle_id": "com.vivaldi.Vivaldi",
        "support_dir": "~/Library/Application Support/Vivaldi",
        "family": "chromium",
    },
    "opera": {
        "label": "Opera",
        "app_paths": [
            "/Applications/Opera.app",
            "~/Applications/Opera.app",
        ],
        "bundle_id": "com.operasoftware.Opera",
        "support_dir": "~/Library/Application Support/com.operasoftware.Opera",
        "family": "chromium",
    },
    "firefox": {
        "label": "Firefox",
        "app_paths": [
            "/Applications/Firefox.app",
            "~/Applications/Firefox.app",
        ],
        "bundle_id": "org.mozilla.firefox",
        "support_dir": "~/Library/Application Support/Firefox/Profiles",
        "family": "firefox",
    },
}
DATE_PATTERNS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d %b %Y",
    "%d %B %Y",
    "%b %d, %Y",
    "%B %d, %Y",
    "%b %d.%Y",
    "%B %d.%Y",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
)


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download AliExpress orders in a date range and export them to CSV."
    )
    parser.add_argument("--start-date", help="Inclusive start date (YYYY-MM-DD).")
    parser.add_argument("--end-date", help="Inclusive end date (YYYY-MM-DD).")
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_EXPORT_DIR,
        help=f"Directory for CSV and PDFs. Default: {DEFAULT_EXPORT_DIR}",
    )
    parser.add_argument(
        "--cookies-path",
        default=DEFAULT_COOKIES_PATH,
        help=f"Path to the stored AliExpress auth cookies. Default: {DEFAULT_COOKIES_PATH}",
    )
    parser.add_argument(
        "--setup",
        action="store_true",
        help="Run an interactive one-time setup that imports cookies from an installed browser.",
    )
    parser.add_argument(
        "--import-har",
        action="append",
        default=[],
        help="Import AliExpress cookies from a HAR file. Can be passed multiple times.",
    )
    parser.add_argument(
        "--input-har",
        action="append",
        default=[],
        help=(
            "Export from HAR capture(s) instead of calling the live API. "
            "Can be passed multiple times."
        ),
    )
    parser.add_argument(
        "--firefox-profile",
        help="Import AliExpress cookies from a Firefox profile directory or cookies.sqlite file.",
    )
    parser.add_argument(
        "--download-pdfs",
        action="store_true",
        help="Download or extract invoice PDFs when available.",
    )
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Also export an XLSX workbook with embedded product thumbnails when available.",
    )
    parser.add_argument(
        "--ship-to-country",
        help="Override ship-to country used for API requests, for example UK.",
    )
    parser.add_argument(
        "--lang",
        help="Override locale used for API requests, for example en_US.",
    )
    parser.add_argument(
        "--time-zone",
        help="Override time zone used for API requests, for example GMT+0000.",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=DEFAULT_MAX_PAGES,
        help=f"Maximum pages to fetch from the AliExpress order list. Default: {DEFAULT_MAX_PAGES}",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=DEFAULT_PAGE_SIZE,
        help=f"Number of orders per API page request. Default: {DEFAULT_PAGE_SIZE}",
    )
    return parser.parse_args(argv)


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def normalize_whitespace(value: str) -> str:
    return " ".join((value or "").split())


def safe_filename(value: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in "._-" else "_" for char in value.strip())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    cleaned = cleaned.strip("._")
    return cleaned or "file"


def parse_human_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        if value > 10_000_000_000:
            value = value / 1000
        try:
            return datetime.fromtimestamp(value, UTC).date()
        except (OverflowError, OSError, ValueError):
            return None
    text = normalize_whitespace(str(value))
    if not text:
        return None
    candidates = [text]
    candidates.extend(re.findall(r"\d{4}-\d{2}-\d{2}", text))
    candidates.extend(re.findall(r"\d{1,2} [A-Za-z]{3,9} \d{4}", text))
    candidates.extend(re.findall(r"[A-Za-z]{3,9} \d{1,2}, \d{4}", text))
    for candidate in candidates:
        normalized = candidate.replace("Sept", "Sep")
        for pattern in DATE_PATTERNS:
            try:
                return datetime.strptime(normalized, pattern).date()
            except ValueError:
                continue
    return None


coerce_order_date = parse_human_date
try_parse_date_from_text = parse_human_date


def default_time_zone() -> str:
    offset = datetime.now().astimezone().strftime("%z")
    return f"GMT{offset}"


def parse_cookie_kv_string(value: str) -> dict[str, str]:
    parts = [part for part in value.split("&") if "=" in part]
    return {key: item for key, item in (part.split("=", 1) for part in parts)}


def sanitize_column_key(key: str) -> str:
    result = []
    for char in key:
        if char.isalnum():
            result.append(char)
        elif char in {"_", "-", "[", "]"}:
            result.append(char)
        else:
            result.append("_")
    return "".join(result).strip("_") or "value"


def flatten_value(prefix: str, value: Any, out: dict[str, str]) -> None:
    key = sanitize_column_key(prefix)
    if isinstance(value, dict):
        if not value:
            out[key] = ""
            return
        for child_key, child_value in sorted(value.items()):
            flatten_value(f"{prefix}.{child_key}", child_value, out)
        return
    if isinstance(value, list):
        if not value:
            out[key] = "[]"
            return
        for index, item in enumerate(value):
            flatten_value(f"{prefix}[{index}]", item, out)
        return
    out[key] = "" if value is None else str(value)


def strip_jsonp_wrapper(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("{") or stripped.startswith("["):
        return stripped
    start = stripped.find("(")
    end = stripped.rfind(")")
    if start != -1 and end != -1 and end > start:
        return stripped[start + 1 : end]
    return stripped


def parse_jsonish_text(text: str) -> dict[str, Any]:
    return json.loads(strip_jsonp_wrapper(text))


def compact_json(value: Any) -> str:
    return json.dumps(value, separators=(",", ":"), ensure_ascii=False)


def ensure_successful_mtop(payload: dict[str, Any], api_name: str) -> dict[str, Any]:
    ret = payload.get("ret") or []
    if ret and not any(str(item).startswith("SUCCESS") for item in ret):
        raise RuntimeError(f"{api_name} failed: {ret}")
    data = payload.get("data")
    if not isinstance(data, dict):
        raise RuntimeError(f"{api_name} returned no data payload.")
    return data


def build_csv_path(output_dir: Path, start: date, end: date) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"aliexpress_orders_{start.isoformat()}_to_{end.isoformat()}.csv"


def build_order_lines_csv_path(output_dir: Path, start: date, end: date) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"aliexpress_order_lines_{start.isoformat()}_to_{end.isoformat()}.csv"


def build_xlsx_path(output_dir: Path, start: date, end: date) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"aliexpress_orders_{start.isoformat()}_to_{end.isoformat()}.xlsx"


def build_pdf_dir(output_dir: Path) -> Path:
    pdf_dir = output_dir / "pdf"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    return pdf_dir


def resolve_order_date(bundle: "OrderBundle") -> Optional[date]:
    candidates = [
        bundle.list_fields.get("orderDateText"),
        bundle.detail_sections.get("detail_simple_order_info_component", {}).get("orderCreatTime"),
    ]
    for candidate in candidates:
        parsed = parse_human_date(candidate)
        if parsed is not None:
            return parsed
    return None


def normalize_detail_section_name(raw_key: str, block: dict[str, Any]) -> str:
    base = block.get("tag") or raw_key
    if raw_key.startswith("detail_order_services_sv3-"):
        return "detail_order_services_sv3"
    while base and base[-1].isdigit():
        base = base[:-1]
    return base.rstrip("_")


def cookie_sort_key(cookie: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(cookie.get("domain", "")),
        str(cookie.get("path", "")),
        str(cookie.get("name", "")),
    )


def sanitize_cookie_value(value: Any) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "")


def dedupe_cookies(cookies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for cookie in cookies:
        sanitized = {
            **cookie,
            "name": str(cookie.get("name", "")),
            "value": sanitize_cookie_value(cookie.get("value", "")),
            "domain": str(cookie.get("domain", "")),
            "path": str(cookie.get("path", "/")),
        }
        key = (
            sanitized["domain"],
            sanitized["path"],
            sanitized["name"],
        )
        deduped[key] = sanitized
    return sorted(deduped.values(), key=cookie_sort_key)


def save_cookies(path: Path, cookies: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(dedupe_cookies(cookies), handle, indent=2, sort_keys=True)


def load_cookies(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        cookies = json.load(handle)
    if not isinstance(cookies, list):
        raise RuntimeError(f"Cookie file {path} is invalid.")
    return dedupe_cookies(cookies)


def cookie_value(cookies: list[dict[str, Any]], name: str) -> str:
    for cookie in cookies:
        if cookie.get("name") == name:
            return str(cookie.get("value", ""))
    return ""


def import_cookies_from_hars(paths: list[Path]) -> list[dict[str, Any]]:
    cookies: list[dict[str, Any]] = []
    for path in paths:
        with path.open("r", encoding="utf-8") as handle:
            entries = json.load(handle).get("log", {}).get("entries", [])
        for entry in entries:
            request_host = urlparse(entry.get("request", {}).get("url", "")).hostname or ""
            for cookie in entry.get("request", {}).get("cookies", []):
                domain = str(cookie.get("domain") or cookie.get("host") or request_host)
                if "aliexpress.com" not in domain:
                    continue
                cookies.append(
                    {
                        "name": cookie.get("name"),
                        "value": cookie.get("value"),
                        "domain": domain or ".aliexpress.com",
                        "path": cookie.get("path") or "/",
                        "expires": cookie.get("expires"),
                        "secure": cookie.get("secure", False),
                        "httpOnly": cookie.get("httpOnly", False),
                    }
                )
    return dedupe_cookies(cookies)


def resolve_firefox_cookie_db(source: Path) -> Path:
    if source.is_file():
        return source
    cookie_db = source / "cookies.sqlite"
    if cookie_db.exists():
        return cookie_db
    raise RuntimeError(f"Could not find cookies.sqlite under {source}")


def import_cookies_from_firefox(source: Path) -> list[dict[str, Any]]:
    cookie_db = resolve_firefox_cookie_db(source)
    with tempfile.TemporaryDirectory() as tmp_dir:
        copied = Path(tmp_dir) / "cookies.sqlite"
        shutil.copy2(cookie_db, copied)
        connection = sqlite3.connect(copied)
        try:
            rows = connection.execute(
                """
                SELECT name, value, host, path, expiry, isSecure, isHttpOnly
                FROM moz_cookies
                WHERE host LIKE '%aliexpress.com%'
                """
            ).fetchall()
        finally:
            connection.close()
    cookies = [
        {
            "name": name,
            "value": value,
            "domain": host,
            "path": path_value,
            "expires": expiry,
            "secure": bool(is_secure),
            "httpOnly": bool(is_http_only),
        }
        for name, value, host, path_value, expiry, is_secure, is_http_only in rows
    ]
    return dedupe_cookies(cookies)


@dataclass
class BrowserProfile:
    browser_id: str
    browser_label: str
    profile_label: str
    profile_path: Path
    cookie_file: Path


def detect_installed_browsers() -> list[str]:
    installed: list[str] = []
    for browser_id, meta in SUPPORTED_BROWSERS.items():
        if find_browser_app_path(meta) is not None:
            installed.append(browser_id)
    return installed


def spotlight_app_paths(bundle_id: str) -> list[Path]:
    try:
        result = subprocess.run(
            ["mdfind", f'kMDItemCFBundleIdentifier == "{bundle_id}"'],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
    except OSError:
        return []
    if result.returncode != 0:
        return []
    paths: list[Path] = []
    for line in result.stdout.splitlines():
        candidate = Path(line.strip()).expanduser()
        if candidate.suffix != ".app":
            continue
        if candidate.exists():
            paths.append(candidate)
    return paths


def find_browser_app_path(meta: dict[str, Any]) -> Optional[Path]:
    for raw_path in meta.get("app_paths", []):
        candidate = Path(raw_path).expanduser()
        if candidate.exists():
            return candidate
    bundle_id = normalize_whitespace(str(meta.get("bundle_id", "")))
    if not bundle_id:
        return None
    spotlight_matches = spotlight_app_paths(bundle_id)
    return spotlight_matches[0] if spotlight_matches else None


def load_chromium_profile_labels(support_dir: Path) -> dict[str, str]:
    local_state = support_dir / "Local State"
    if not local_state.exists():
        return {}
    try:
        data = json.loads(local_state.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    info_cache = data.get("profile", {}).get("info_cache", {})
    labels: dict[str, str] = {}
    if not isinstance(info_cache, dict):
        return labels
    for profile_dir, profile_meta in info_cache.items():
        if not isinstance(profile_meta, dict):
            continue
        display_name = normalize_whitespace(str(profile_meta.get("name", "")))
        if not display_name:
            continue
        labels[str(profile_dir)] = display_name
    return labels


def load_firefox_profile_labels(support_dir: Path) -> dict[str, str]:
    profiles_ini = support_dir.parent / "profiles.ini"
    if not profiles_ini.exists():
        return {}
    parser = configparser.ConfigParser()
    try:
        parser.read(profiles_ini, encoding="utf-8")
    except (OSError, configparser.Error):
        return {}
    labels: dict[str, str] = {}
    for section in parser.sections():
        if not section.startswith("Profile"):
            continue
        path_value = normalize_whitespace(parser.get(section, "Path", fallback=""))
        if not path_value:
            continue
        name = normalize_whitespace(parser.get(section, "Name", fallback=""))
        if name:
            labels[Path(path_value).name] = name
    return labels


def firefox_profile_fallback_label(profile_dir_name: str) -> str:
    match = re.match(r"^[A-Za-z0-9]{8}\.(.+)$", profile_dir_name)
    if match:
        return normalize_whitespace(match.group(1))
    return profile_dir_name


def uniquify_profile_labels(profiles: list[BrowserProfile]) -> list[BrowserProfile]:
    counts: dict[str, int] = {}
    for profile in profiles:
        counts[profile.profile_label] = counts.get(profile.profile_label, 0) + 1

    updated: list[BrowserProfile] = []
    for profile in profiles:
        profile_label = profile.profile_label
        if counts.get(profile_label, 0) > 1:
            profile_label = f"{profile_label} ({profile.profile_path.name})"
        updated.append(
            BrowserProfile(
                browser_id=profile.browser_id,
                browser_label=profile.browser_label,
                profile_label=profile_label,
                profile_path=profile.profile_path,
                cookie_file=profile.cookie_file,
            )
        )
    return updated


def detect_chromium_profiles(
    browser_id: str, browser_label: str, support_dir: Path
) -> list[BrowserProfile]:
    profiles: list[BrowserProfile] = []
    if not support_dir.exists():
        return profiles
    profile_labels = load_chromium_profile_labels(support_dir)
    candidates = []
    for child in sorted(support_dir.iterdir()):
        if not child.is_dir():
            continue
        if child.name in {"System Profile", "Guest Profile"}:
            continue
        cookie_file = child / "Network" / "Cookies"
        if not cookie_file.exists():
            cookie_file = child / "Cookies"
        if cookie_file.exists():
            candidates.append((child.name, child, cookie_file))
    for profile_name, profile_path, cookie_file in candidates:
        display_name = profile_labels.get(profile_name, profile_name)
        profiles.append(
            BrowserProfile(
                browser_id=browser_id,
                browser_label=browser_label,
                profile_label=display_name,
                profile_path=profile_path,
                cookie_file=cookie_file,
            )
        )
    return uniquify_profile_labels(profiles)


def detect_firefox_profiles(
    browser_id: str, browser_label: str, support_dir: Path
) -> list[BrowserProfile]:
    profiles: list[BrowserProfile] = []
    if not support_dir.exists():
        return profiles
    profile_labels = load_firefox_profile_labels(support_dir)
    for child in sorted(support_dir.iterdir()):
        cookie_file = child / "cookies.sqlite"
        if child.is_dir() and cookie_file.exists():
            fallback_name = firefox_profile_fallback_label(child.name)
            display_name = profile_labels.get(child.name, fallback_name)
            profiles.append(
                BrowserProfile(
                    browser_id=browser_id,
                    browser_label=browser_label,
                    profile_label=display_name,
                    profile_path=child,
                    cookie_file=cookie_file,
                )
            )
    return uniquify_profile_labels(profiles)


def detect_installed_browser_profiles(
    browser_ids: Optional[list[str]] = None,
) -> list[BrowserProfile]:
    profiles: list[BrowserProfile] = []
    selected_browser_ids = browser_ids or list(SUPPORTED_BROWSERS)
    for browser_id in selected_browser_ids:
        meta = SUPPORTED_BROWSERS[browser_id]
        app_path = find_browser_app_path(meta)
        support_dir = Path(meta["support_dir"]).expanduser()
        if app_path is None or not support_dir.exists():
            continue
        if meta["family"] == "firefox":
            profiles.extend(
                detect_firefox_profiles(browser_id, meta["label"], support_dir)
            )
        else:
            profiles.extend(
                detect_chromium_profiles(browser_id, meta["label"], support_dir)
            )
    return profiles


def group_profiles_by_browser(profiles: list[BrowserProfile]) -> dict[str, list[BrowserProfile]]:
    grouped: dict[str, list[BrowserProfile]] = {}
    for profile in profiles:
        grouped.setdefault(profile.browser_id, []).append(profile)
    return grouped


def prompt_index(prompt: str, max_value: int) -> int:
    while True:
        raw = input(prompt).strip()
        if raw.lower() in {"q", "quit", "exit"}:
            raise SystemExit("Setup cancelled.")
        if raw.isdigit():
            choice = int(raw)
            if 1 <= choice <= max_value:
                return choice - 1
        print(f"Enter a number between 1 and {max_value}, or 'q' to cancel.")


def prompt_enter_to_continue() -> None:
    confirm = input().strip().lower()
    if confirm in {"q", "quit", "exit"}:
        raise SystemExit("Setup cancelled.")


def load_browser_cookie3() -> Any:
    try:
        import browser_cookie3
    except ImportError as exc:
        raise SystemExit(
            "Interactive browser setup requires browser-cookie3.\n"
            "Run with uv so dependencies resolve, for example:\n"
            "  uv run aliexpress_export.py --setup\n"
        ) from exc
    return browser_cookie3


def cookiejar_to_dicts(cookiejar: Any) -> list[dict[str, Any]]:
    cookies: list[dict[str, Any]] = []
    for cookie in cookiejar:
        if "aliexpress.com" not in cookie.domain:
            continue
        cookies.append(
            {
                "name": cookie.name,
                "value": cookie.value,
                "domain": cookie.domain,
                "path": cookie.path,
                "expires": cookie.expires,
                "secure": cookie.secure,
                "httpOnly": bool(cookie._rest.get("HttpOnly")),  # noqa: SLF001
            }
        )
    return dedupe_cookies(cookies)


def import_cookies_from_browser_profile(profile: BrowserProfile) -> list[dict[str, Any]]:
    browser_cookie3 = load_browser_cookie3()
    loader = getattr(browser_cookie3, profile.browser_id)
    cookiejar = loader(cookie_file=str(profile.cookie_file), domain_name="aliexpress.com")
    cookies = cookiejar_to_dicts(cookiejar)
    if not cookies:
        raise RuntimeError(
            "No AliExpress cookies were found in "
            f"{profile.browser_label} profile {profile.profile_label}."
        )
    return cookies


def test_live_connection(cookies_path: Path) -> tuple[bool, str]:
    client = AliExpressMtopClient(load_cookies(cookies_path))
    try:
        client.bootstrap()
        payload = client.fetch_order_list_page(1, DEFAULT_PAGE_SIZE)
        store = OrderStore()
        parse_order_list_payload(payload, store)
        return True, f"Connected successfully. Found {len(store.orders)} orders on page 1."
    except Exception as exc:
        return False, str(exc)
    finally:
        client.close()


def run_interactive_setup(config: argparse.Namespace) -> int:
    cookies_path = Path(config.cookies_path).expanduser().resolve()
    installed_browser_ids = detect_installed_browsers()
    if not installed_browser_ids:
        raise SystemExit(
            "No supported browsers were found.\n"
            "Supported browsers: Chrome, Brave, Edge, Chromium, Vivaldi, Opera, Firefox."
        )

    browser_ids = sorted(installed_browser_ids, key=lambda item: SUPPORTED_BROWSERS[item]["label"])
    print("Select the browser you use for AliExpress:\n")
    for index, browser_id in enumerate(browser_ids, start=1):
        print(f"{index}. {SUPPORTED_BROWSERS[browser_id]['label']}")
    browser_index = prompt_index("\nBrowser: ", len(browser_ids))
    selected_browser_id = browser_ids[browser_index]
    browser_profiles = detect_installed_browser_profiles([selected_browser_id])
    if not browser_profiles:
        raise SystemExit(
            "No readable profiles were found for "
            f"{SUPPORTED_BROWSERS[selected_browser_id]['label']}.\n"
            "Open that browser once with the profile you use for AliExpress, then run setup again."
        )

    print(
        f"\nUsing {SUPPORTED_BROWSERS[selected_browser_id]['label']}.\n"
        "Make sure you are logged into AliExpress in that browser now.\n"
        "Press Enter when ready, or type 'q' to cancel."
    )
    prompt_enter_to_continue()

    grouped = group_profiles_by_browser(browser_profiles)
    browser_profiles = grouped[selected_browser_id]

    selected_profile = browser_profiles[0]
    if len(browser_profiles) > 1:
        print(f"\nProfiles found for {selected_profile.browser_label}:\n")
        for index, profile in enumerate(browser_profiles, start=1):
            print(f"{index}. {profile.profile_label}")
        profile_index = prompt_index("\nProfile: ", len(browser_profiles))
        selected_profile = browser_profiles[profile_index]

    print(
        f"\nSelected {selected_profile.browser_label} profile '{selected_profile.profile_label}'.\n"
        "Press Enter to import cookies and test the live connection, or type 'q' to cancel."
    )
    prompt_enter_to_continue()

    cookies = import_cookies_from_browser_profile(selected_profile)
    save_cookies(cookies_path, cookies)
    print(f"\nImported {len(cookies)} AliExpress cookies into {cookies_path}")
    print("Testing live connection...")
    ok, message = test_live_connection(cookies_path)
    if not ok:
        raise SystemExit(
            "Cookie import succeeded, but the live connection test failed.\n"
            f"{message}\n"
            "Check that this browser profile is still logged into AliExpress and try setup again."
        )
    print(message)
    return 0


@dataclass
class OrderBundle:
    order_id: str
    list_fields: dict[str, Any] = field(default_factory=dict)
    detail_sections: dict[str, dict[str, Any]] = field(default_factory=dict)
    invoice_info_list: list[dict[str, Any]] = field(default_factory=list)
    invoice_files: list[dict[str, Any]] = field(default_factory=list)
    invoice_pdf_paths: list[str] = field(default_factory=list)

    def order_date(self) -> str:
        parsed = resolve_order_date(self)
        return parsed.isoformat() if parsed else ""

    def base_row(self) -> dict[str, str]:
        total_price = self.list_fields.get("totalPriceText") or (
            self.detail_sections.get("detail_order_price_block", {})
            .get("totalPrice", {})
            .get("value", "")
        )
        list_lines = self.list_fields.get("orderLines") or []
        detail_lines = (
            self.detail_sections.get("detail_product_block", {}).get("productVOList") or []
        )
        lines = detail_lines or list_lines
        titles = [normalize_whitespace(str(line.get("itemTitle", ""))) for line in lines]
        titles = [title for title in titles if title]
        image_urls = [str(line.get("itemImgUrl", "")).strip() for line in lines]
        image_urls = [url for url in image_urls if url]
        simple_info = self.detail_sections.get("detail_simple_order_info_component", {})
        address = simple_info.get("addressVO", {})
        invoice_url = ""
        status_buttons = self.detail_sections.get("detail_order_status_block", {}).get(
            "buttonVOList", []
        )
        for button in status_buttons:
            if button.get("type") == "DOWNLOAD_INVOICE":
                invoice_url = str(button.get("href", ""))
                break
        return {
            "order_id": self.order_id,
            "order_date": self.order_date(),
            "status": str(self.list_fields.get("statusText", "")),
            "total_price_text": str(total_price),
            "currency": str(self.list_fields.get("currencyCode", "")),
            "store_name": str(self.list_fields.get("storeName", "")),
            "order_title": titles[0] if titles else "",
            "order_titles": " | ".join(titles),
            "item_image_url": image_urls[0] if image_urls else "",
            "order_line_count": str(len(lines)),
            "payment_method": str(simple_info.get("paymentMethod", "")),
            "shipping_contact_name": str(address.get("contactName", "")),
            "shipping_country_code": str(address.get("countryCode", "")),
            "shipping_post_code": str(address.get("postCode", "")),
            "shipping_region_address": str(address.get("regionAddress", "")),
            "shipping_detail_address": str(address.get("detailAddress", "")),
            "shipping_phone": str(address.get("fullPhoneNo", "")),
            "order_detail_url": str(self.list_fields.get("orderDetailUrl", "")),
            "invoice_page_url": invoice_url,
            "invoice_pdf_paths": json.dumps(self.invoice_pdf_paths, ensure_ascii=False),
            "list_json": json.dumps(self.list_fields, ensure_ascii=False),
            "detail_json": json.dumps(self.detail_sections, ensure_ascii=False),
            "invoice_info_json": json.dumps(self.invoice_info_list, ensure_ascii=False),
        }

    def order_line_rows(self) -> list[dict[str, str]]:
        detail_lines = (
            self.detail_sections.get("detail_product_block", {}).get("productVOList") or []
        )
        list_lines = self.list_fields.get("orderLines") or []
        rows: list[dict[str, str]] = []
        for index, line in enumerate(detail_lines or list_lines):
            rows.append(
                {
                    "order_id": self.order_id,
                    "order_date": self.order_date(),
                    "store_name": str(self.list_fields.get("storeName", "")),
                    "status": str(self.list_fields.get("statusText", "")),
                    "line_index": str(index),
                    "order_line_id": str(line.get("orderLineId", "")),
                    "product_id": str(line.get("productId", "")),
                    "sku_id": str(line.get("skuId", "")),
                    "item_title": str(line.get("itemTitle", "")),
                    "item_image_url": str(line.get("itemImgUrl", "")),
                    "item_price_text": str(line.get("itemPriceText", "")),
                    "currency": str(
                        line.get("currencyCode", self.list_fields.get("currencyCode", ""))
                    ),
                    "quantity": str(line.get("quantity", "")),
                    "item_detail_url": str(line.get("itemDetailUrl", "")),
                    "snapshot_url": str(line.get("snapshotUrl", "")),
                    "sku_attrs_json": json.dumps(line.get("skuAttrs", []), ensure_ascii=False),
                    "item_tags_json": json.dumps(line.get("itemTags", []), ensure_ascii=False),
                    "line_json": json.dumps(line, ensure_ascii=False),
                }
            )
        return rows


class OrderStore:
    def __init__(self) -> None:
        self.orders: dict[str, OrderBundle] = {}

    def get(self, order_id: str) -> OrderBundle:
        if order_id not in self.orders:
            self.orders[order_id] = OrderBundle(order_id=order_id)
        return self.orders[order_id]

    def add_list_fields(self, order_id: str, fields: dict[str, Any]) -> None:
        self.get(order_id).list_fields.update(fields)

    def add_detail_sections(self, order_id: str, sections: dict[str, dict[str, Any]]) -> None:
        self.get(order_id).detail_sections.update(sections)

    def add_invoice_info(self, order_id: str, invoice_info_list: list[dict[str, Any]]) -> None:
        self.get(order_id).invoice_info_list = invoice_info_list

    def add_invoice_file(self, order_id: str, invoice_file: dict[str, Any]) -> None:
        self.get(order_id).invoice_files.append(invoice_file)

    def add_invoice_pdf_path(self, order_id: str, pdf_path: Path) -> None:
        bundle = self.get(order_id)
        path_str = str(pdf_path)
        if path_str not in bundle.invoice_pdf_paths:
            bundle.invoice_pdf_paths.append(path_str)

    def filtered(self, start: date, end: date) -> list[OrderBundle]:
        bundles = []
        for bundle in self.orders.values():
            parsed = resolve_order_date(bundle)
            if parsed is None:
                continue
            if start <= parsed <= end:
                bundles.append(bundle)
        bundles.sort(key=lambda bundle: (bundle.order_date(), bundle.order_id), reverse=True)
        return bundles


def parse_order_list_payload(payload: dict[str, Any], store: OrderStore) -> dict[str, Any]:
    data = ensure_successful_mtop(payload, "order.list")
    blocks = data.get("data", {})
    metadata: dict[str, Any] = {}
    for key, block in blocks.items():
        if not isinstance(block, dict):
            continue
        fields = block.get("fields")
        if not isinstance(fields, dict):
            continue
        tag = block.get("tag") or key
        if tag == "pc_om_list_order" or key.startswith("pc_om_list_order_"):
            order_id = str(fields.get("orderId", "")).strip()
            if order_id:
                store.add_list_fields(order_id, fields)
        elif tag == "pc_om_list_body":
            metadata = fields
    return metadata


def parse_optional_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = normalize_whitespace(str(value))
    if not text:
        return None
    if text.isdigit():
        return int(text)
    return None


def find_block_name_by_tag(data_blocks: dict[str, Any], tag: str) -> Optional[str]:
    for block_name, block in data_blocks.items():
        if isinstance(block, dict) and block.get("tag") == tag:
            return block_name
    return None


def repeated_pagination_reason(
    requested_page_index: int,
    new_orders: int,
    reported_page_index: Optional[int],
    last_reported_page_index: Optional[int],
) -> Optional[str]:
    if requested_page_index > 1 and new_orders <= 0:
        return "AliExpress returned no new orders."
    if (
        reported_page_index is not None
        and last_reported_page_index is not None
        and reported_page_index <= last_reported_page_index
    ):
        return f"AliExpress reported pageIndex={reported_page_index} again."
    return None


def parse_order_detail_payload(payload: dict[str, Any], store: OrderStore) -> str:
    data = ensure_successful_mtop(payload, "order.detail")
    blocks = data.get("data", {})
    order_id = ""
    sections: dict[str, dict[str, Any]] = {}
    for key, block in blocks.items():
        if not isinstance(block, dict):
            continue
        fields = block.get("fields")
        if not isinstance(fields, dict):
            continue
        section_name = normalize_detail_section_name(key, block)
        sections[section_name] = fields
        if not order_id:
            candidate = fields.get("orderId")
            if candidate is None:
                candidate = fields.get("tradeOrderId")
            if candidate is not None:
                order_id = str(candidate)
    if not order_id:
        raise RuntimeError("Could not determine order id from order detail payload.")
    store.add_detail_sections(order_id, sections)
    return order_id


def parse_invoice_info_payload(payload: dict[str, Any], order_id: str, store: OrderStore) -> None:
    data = ensure_successful_mtop(payload, "invoice.info")
    invoice_info_list = data.get("data", {}).get("invoiceInfoDTOList", [])
    if isinstance(invoice_info_list, list):
        store.add_invoice_info(order_id, invoice_info_list)


def write_orders_csv(path: Path, bundles: list[OrderBundle]) -> None:
    rows = [bundle.base_row() for bundle in bundles]
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_order_lines_csv(path: Path, bundles: list[OrderBundle]) -> None:
    rows = [row for bundle in bundles for row in bundle.order_line_rows()]
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def download_image_file(url: str, target_path: Path) -> Optional[Path]:
    if not url:
        return None
    result = subprocess.run(
        [
            "curl",
            "--silent",
            "--show-error",
            "--location",
            "--compressed",
            "--output",
            str(target_path),
            url,
        ],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        env=os.environ.copy(),
    )
    if result.returncode != 0 or not target_path.exists() or target_path.stat().st_size == 0:
        return None
    return target_path


def configure_sheet_columns(sheet: Any, widths: list[tuple[str, float]]) -> None:
    for column_name, width in widths:
        sheet.column_dimensions[column_name].width = width


def style_header_row(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def write_sheet(
    sheet: Any,
    columns: list[tuple[str, str]],
    rows: list[dict[str, str]],
    thumbnail_key: str,
    image_cache: dict[str, Optional[Path]],
    temp_dir: Path,
) -> None:
    headers = [header for header, _key in columns]
    sheet.append(headers)
    style_header_row(sheet)
    thumbnail_col_idx = headers.index("Thumbnail") + 1 if "Thumbnail" in headers else 0
    for row_idx, row in enumerate(rows, start=2):
        values = []
        image_url = ""
        for header, key in columns:
            if key == "__thumbnail__":
                values.append("")
                image_url = row.get(thumbnail_key, "")
            else:
                values.append(row.get(key, ""))
        sheet.append(values)
        if not image_url or not thumbnail_col_idx:
            continue
        if image_url not in image_cache:
            suffix = Path(urlparse(image_url).path).suffix or ".jpg"
            filename = safe_filename(hashlib.md5(image_url.encode("utf-8")).hexdigest())
            target_path = temp_dir / f"{filename}{suffix}"
            image_cache[image_url] = download_image_file(image_url, target_path)
        image_path = image_cache.get(image_url)
        if image_path is None:
            continue
        try:
            image = OpenPyxlImage(str(image_path))
        except Exception:
            continue
        image.width = 72
        image.height = 72
        anchor = f"{get_column_letter(thumbnail_col_idx)}{row_idx}"
        sheet.add_image(image, anchor)
        sheet.row_dimensions[row_idx].height = 56
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_xlsx(path: Path, bundles: list[OrderBundle]) -> None:
    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Orders"
    lines_sheet = workbook.create_sheet("Order Lines")

    order_rows = [bundle.base_row() for bundle in bundles]
    line_rows = [row for bundle in bundles for row in bundle.order_line_rows()]

    order_columns = [
        ("Order ID", "order_id"),
        ("Order Date", "order_date"),
        ("Thumbnail", "__thumbnail__"),
        ("Order Title", "order_title"),
        ("All Titles", "order_titles"),
        ("Status", "status"),
        ("Total", "total_price_text"),
        ("Currency", "currency"),
        ("Store Name", "store_name"),
        ("Payment Method", "payment_method"),
        ("Shipping Contact", "shipping_contact_name"),
        ("Shipping Country", "shipping_country_code"),
        ("Post Code", "shipping_post_code"),
        ("Region Address", "shipping_region_address"),
        ("Detail Address", "shipping_detail_address"),
        ("Phone", "shipping_phone"),
        ("Order Detail URL", "order_detail_url"),
        ("Invoice Page URL", "invoice_page_url"),
        ("Invoice PDFs", "invoice_pdf_paths"),
        ("List JSON", "list_json"),
        ("Detail JSON", "detail_json"),
        ("Invoice JSON", "invoice_info_json"),
    ]
    line_columns = [
        ("Order ID", "order_id"),
        ("Order Date", "order_date"),
        ("Thumbnail", "__thumbnail__"),
        ("Item Title", "item_title"),
        ("Store Name", "store_name"),
        ("Status", "status"),
        ("Price", "item_price_text"),
        ("Currency", "currency"),
        ("Quantity", "quantity"),
        ("Order Line ID", "order_line_id"),
        ("Product ID", "product_id"),
        ("SKU ID", "sku_id"),
        ("Item Detail URL", "item_detail_url"),
        ("Snapshot URL", "snapshot_url"),
        ("SKU Attrs JSON", "sku_attrs_json"),
        ("Item Tags JSON", "item_tags_json"),
        ("Line JSON", "line_json"),
    ]

    with tempfile.TemporaryDirectory(prefix="aliexpress-xlsx-images-") as tmp_dir:
        temp_dir = Path(tmp_dir)
        image_cache: dict[str, Optional[Path]] = {}
        write_sheet(
            orders_sheet,
            order_columns,
            order_rows,
            thumbnail_key="item_image_url",
            image_cache=image_cache,
            temp_dir=temp_dir,
        )
        write_sheet(
            lines_sheet,
            line_columns,
            line_rows,
            thumbnail_key="item_image_url",
            image_cache=image_cache,
            temp_dir=temp_dir,
        )
        configure_sheet_columns(
            orders_sheet,
            [
                ("A", 18),
                ("B", 14),
                ("C", 12),
                ("D", 48),
                ("E", 60),
                ("F", 14),
                ("G", 12),
                ("H", 10),
                ("I", 28),
                ("J", 16),
                ("K", 18),
                ("L", 14),
                ("M", 12),
                ("N", 24),
                ("O", 28),
                ("P", 18),
                ("Q", 28),
                ("R", 28),
                ("S", 28),
                ("T", 40),
                ("U", 40),
                ("V", 32),
            ],
        )
        configure_sheet_columns(
            lines_sheet,
            [
                ("A", 18),
                ("B", 14),
                ("C", 12),
                ("D", 48),
                ("E", 28),
                ("F", 14),
                ("G", 12),
                ("H", 10),
                ("I", 10),
                ("J", 18),
                ("K", 16),
                ("L", 18),
                ("M", 28),
                ("N", 28),
                ("O", 28),
                ("P", 28),
                ("Q", 36),
            ],
        )
        workbook.save(path)


class AliExpressMtopClient:
    def __init__(
        self,
        cookies: list[dict[str, Any]],
        ship_to_country: Optional[str] = None,
        lang: Optional[str] = None,
        time_zone: Optional[str] = None,
    ) -> None:
        self.cookies = dedupe_cookies(cookies)
        self.cookie_dir = Path(tempfile.mkdtemp(prefix="aliexpress-cookies-"))
        self.cookie_file = self.cookie_dir / "cookies.txt"
        self.acs_ip: Optional[str] = None
        self._write_cookie_file()
        self.ship_to_country = ship_to_country or self._cookie_context("region") or "US"
        if self.ship_to_country == "CN":
            self.ship_to_country = "US"
        self.lang = lang or self._cookie_context("b_locale") or "en_US"
        self.time_zone = time_zone or default_time_zone()

    def close(self) -> None:
        shutil.rmtree(self.cookie_dir, ignore_errors=True)

    def export_cookies(self) -> list[dict[str, Any]]:
        return self.cookies

    def bootstrap(self) -> None:
        self._run_curl([DEFAULT_ORDERS_PAGE_URL], expect_json=False)

    def _write_cookie_file(self) -> None:
        lines = ["# Netscape HTTP Cookie File"]
        for cookie in self.cookies:
            domain = str(cookie.get("domain") or ".aliexpress.com")
            include_subdomains = "TRUE" if domain.startswith(".") else "FALSE"
            path_value = str(cookie.get("path") or "/")
            secure = "TRUE" if cookie.get("secure") else "FALSE"
            expires = str(int(cookie.get("expires") or 0))
            name = str(cookie.get("name") or "")
            value = sanitize_cookie_value(cookie.get("value", ""))
            if not name:
                continue
            prefix = "#HttpOnly_" if cookie.get("httpOnly") else ""
            lines.append(
                "\t".join(
                    [
                        f"{prefix}{domain}",
                        include_subdomains,
                        path_value,
                        secure,
                        expires,
                        name,
                        value,
                    ]
                )
            )
        self.cookie_file.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def _reload_cookies_from_file(self) -> None:
        cookies: list[dict[str, Any]] = []
        if not self.cookie_file.exists():
            return
        for line in self.cookie_file.read_text(encoding="utf-8").splitlines():
            if not line or line.startswith("# Netscape HTTP Cookie File"):
                continue
            http_only = False
            if line.startswith("#HttpOnly_"):
                http_only = True
                line = line[len("#HttpOnly_") :]
            if line.startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) != 7:
                continue
            domain, _include_subdomains, path_value, secure, expires, name, value = parts
            cookies.append(
                {
                    "name": name,
                    "value": value,
                    "domain": domain,
                    "path": path_value,
                    "expires": int(expires or 0),
                    "secure": secure == "TRUE",
                    "httpOnly": http_only,
                }
            )
        self.cookies = dedupe_cookies(cookies)

    def _cookie_context(self, key: str) -> str:
        raw = ""
        for name in ("aep_usuc_f", "xman_us_f"):
            raw = cookie_value(self.cookies, name)
            if raw:
                values = parse_cookie_kv_string(raw)
                if key in values:
                    return values[key]
        return ""

    def _token(self) -> str:
        raw = ""
        candidates = []
        for cookie in self.cookies:
            if cookie["name"] != "_m_h5_tk":
                continue
            candidates.append(cookie)
        candidates.sort(key=lambda cookie: 0 if "aliexpress.com" in cookie["domain"] else 1)
        if candidates:
            raw = str(candidates[0]["value"])
        if not raw:
            raise RuntimeError("AliExpress auth token (_m_h5_tk) is missing from stored cookies.")
        return raw.split("_", 1)[0]

    def _run_curl(self, args: list[str], *, expect_json: bool) -> dict[str, Any]:
        env = os.environ.copy()
        target_url = next(
            (arg for arg in args if arg.startswith("https://acs.aliexpress.com/")),
            "",
        )
        for attempt in range(5):
            cmd = [
                "curl",
                "--silent",
                "--show-error",
                "--location",
                "--compressed",
                "--cookie",
                str(self.cookie_file),
                "--cookie-jar",
                str(self.cookie_file),
            ]
            if target_url:
                ip = self._select_acs_ip(force_refresh=attempt > 0)
                if ip:
                    cmd.extend(["--resolve", f"acs.aliexpress.com:443:{ip}"])
            cmd.extend(args)
            result = subprocess.run(
                cmd,
                check=False,
                capture_output=True,
                text=True,
                encoding="utf-8",
                env=env,
            )
            if result.returncode == 0:
                self._reload_cookies_from_file()
                return parse_jsonish_text(result.stdout) if expect_json else {}
            if "SSL_connect" not in result.stderr and "SSL_ERROR_SYSCALL" not in result.stderr:
                message = result.stderr.strip() or f"curl failed with exit code {result.returncode}"
                raise RuntimeError(message)
            self.acs_ip = None
        message = result.stderr.strip() or f"curl failed with exit code {result.returncode}"
        raise RuntimeError(message)

    def _select_acs_ip(self, force_refresh: bool = False) -> str:
        if self.acs_ip and not force_refresh:
            return self.acs_ip
        probe = [
            "curl",
            "--silent",
            "--show-error",
            "--location",
            "--output",
            "/dev/null",
            "--write-out",
            "%{remote_ip}",
            "https://acs.aliexpress.com/",
        ]
        result = subprocess.run(
            probe,
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            env=os.environ.copy(),
        )
        self.acs_ip = result.stdout.strip() if result.returncode == 0 else ""
        return self.acs_ip

    def _sign(self, timestamp: str, data: str, app_key: str = DEFAULT_API_APP_KEY) -> str:
        raw = f"{self._token()}&{timestamp}&{app_key}&{data}"
        return hashlib.md5(raw.encode("utf-8")).hexdigest()

    def _request(
        self,
        api_name: str,
        data: dict[str, Any],
        *,
        method: str,
        path_api_name: Optional[str] = None,
        app_key: str = DEFAULT_API_APP_KEY,
        data_type: str,
        request_type: str,
        callback: Optional[str] = None,
        need_login: bool = False,
        timeout_ms: Optional[int] = None,
        ecode: bool = False,
        extra_query_params: Optional[dict[str, Any]] = None,
    ) -> dict[str, Any]:
        data_str = compact_json(data)
        timestamp = str(int(time.time() * 1000))
        params: dict[str, Any] = {
            "appKey": app_key,
            "t": timestamp,
            "sign": self._sign(timestamp, data_str, app_key),
            "api": api_name,
            "v": "1.0",
            "type": request_type,
            "dataType": data_type,
        }
        if method == "GET":
            params["jsv"] = "2.5.1"
            params["method"] = "GET"
            params["timeout"] = str(timeout_ms or 15000)
            params["data"] = data_str
            if callback:
                params["callback"] = callback
        else:
            if need_login:
                params["needLogin"] = "true"
            if ecode:
                params["ecode"] = "1"
        if extra_query_params:
            params.update(extra_query_params)
        endpoint_name = (path_api_name or api_name).lower()
        url = f"https://acs.aliexpress.com/h5/{endpoint_name}/1.0/"
        headers = {
            "origin": "https://www.aliexpress.com",
            "referer": DEFAULT_ORDERS_PAGE_URL,
            "user-agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
            ),
        }
        query = urlencode(params)
        curl_args = [
            "--header",
            f"origin: {headers['origin']}",
            "--header",
            f"referer: {headers['referer']}",
            "--header",
            f"user-agent: {headers['user-agent']}",
        ]
        if method == "GET":
            curl_args.append(f"{url}?{query}")
        else:
            curl_args.extend(
                [
                    "--request",
                    "POST",
                    "--header",
                    "content-type: application/x-www-form-urlencoded",
                    "--data-urlencode",
                    f"data={data_str}",
                    f"{url}?{query}",
                ]
            )
        return self._run_curl(curl_args, expect_json=True)

    def fetch_order_list_page(self, page_index: int, page_size: int) -> dict[str, Any]:
        data: dict[str, Any] = {
            "statusTab": None,
            "renderType": "init",
            "clientPlatform": "pc",
            "shipToCountry": self.ship_to_country,
            "_lang": self.lang,
            "timeZone": self.time_zone,
        }
        return self._request(
            "mtop.aliexpress.trade.buyer.order.list",
            data,
            method="GET",
            data_type="originaljsonp",
            request_type="originaljsonp",
            callback="mtopjsonp1",
        )

    def fetch_order_list_page_more(
        self,
        previous_payload: dict[str, Any],
        page_index: int,
        page_size: int,
    ) -> dict[str, Any]:
        payload_data = previous_payload.get("data", {})
        if not isinstance(payload_data, dict):
            raise RuntimeError("Previous order list payload is missing data.")
        data_blocks = payload_data.get("data", {})
        linkage = payload_data.get("linkage")
        hierarchy = payload_data.get("hierarchy")
        endpoint = payload_data.get("endpoint")
        if not isinstance(data_blocks, dict):
            raise RuntimeError("Previous order list payload is missing block data.")
        body_block_name = find_block_name_by_tag(data_blocks, "pc_om_list_body")
        header_action_name = find_block_name_by_tag(data_blocks, "pc_om_list_header_action")
        if (
            body_block_name is None
            or header_action_name is None
            or not isinstance(linkage, dict)
            or not isinstance(hierarchy, dict)
            or not isinstance(endpoint, dict)
        ):
            raise RuntimeError("Previous order list payload is missing pagination state.")

        request_blocks = {
            body_block_name: json.loads(compact_json(data_blocks[body_block_name])),
            header_action_name: json.loads(compact_json(data_blocks[header_action_name])),
        }
        body_fields = request_blocks[body_block_name].setdefault("fields", {})
        body_fields["pageIndex"] = page_index
        body_fields["pageSize"] = page_size

        post_payload = {
            "params": compact_json(
                {
                    "data": compact_json(request_blocks),
                    "linkage": compact_json(linkage),
                    "hierarchy": compact_json(hierarchy),
                    "endpoint": compact_json(endpoint),
                    "operator": body_block_name,
                }
            ),
            "shipToCountry": self.ship_to_country,
            "_lang": self.lang,
        }
        return self._request(
            "mtop.aliexpress.trade.buyer.order.list",
            post_payload,
            method="POST",
            data_type="originaljsonp",
            request_type="originaljson",
            need_login=True,
            ecode=True,
            extra_query_params={"post": "1", "isSec": "1"},
        )

    def fetch_order_detail(self, order_id: str) -> dict[str, Any]:
        return self._request(
            "mtop.aliexpress.trade.buyer.order.detail",
            {
                "tradeOrderId": order_id,
                "clientPlatform": "pc",
                "shipToCountry": self.ship_to_country,
                "_lang": self.lang,
                "timeZone": self.time_zone,
            },
            method="GET",
            data_type="originaljsonp",
            request_type="originaljsonp",
            callback="mtopjsonp1",
        )

    def fetch_invoice_info(self, order_id: str) -> dict[str, Any]:
        return self._request(
            "mtop.global.finance.taxation.invoice.getInvoiceInfoListByType",
            {
                "orderId": order_id,
                "shipToCountry": self.ship_to_country,
                "_lang": self.lang,
                "timeZone": self.time_zone,
                "clientPlatform": "pc",
                "invoiceType": "INVOICE",
            },
            method="POST",
            path_api_name="mtop.global.finance.taxation.invoice.getinvoiceinfolistbytype",
            data_type="jsonp",
            request_type="originaljson",
            need_login=True,
            ecode=True,
        )

    def fetch_invoice_file(self, order_id: str, invoice: dict[str, Any]) -> bytes:
        payload = self._request(
            "mtop.global.finance.taxation.invoice.getInvoiceFileContent",
            {
                "orderId": order_id,
                "shipToCountry": self.ship_to_country,
                "_lang": self.lang,
                "timeZone": self.time_zone,
                "clientPlatform": "pc",
                "invoiceId": invoice.get("invoiceId"),
                "invoiceNo": invoice.get("invoiceNo"),
            },
            method="POST",
            path_api_name="mtop.global.finance.taxation.invoice.getinvoicefilecontent",
            data_type="jsonp",
            request_type="originaljson",
            need_login=True,
            ecode=True,
        )
        data = ensure_successful_mtop(payload, "invoice.file")
        content = data.get("data", {}).get("content", "")
        if not content:
            raise RuntimeError(f"No invoice PDF content returned for order {order_id}")
        return base64.b64decode(content)


def invoice_pdf_path(pdf_dir: Path, order_id: str, invoice: dict[str, Any], index: int) -> Path:
    invoice_id = str(invoice.get("invoiceId") or f"idx{index + 1}")
    invoice_no = str(invoice.get("invoiceNo") or "")
    name = f"invoice_{safe_filename(order_id)}_{safe_filename(invoice_id)}"
    if invoice_no:
        name += f"_{safe_filename(invoice_no)}"
    return pdf_dir / f"{name}.pdf"


def maybe_write_pdf(path: Path, content: bytes) -> bool:
    if path.exists():
        return False
    path.write_bytes(content)
    return True


def export_live(config: argparse.Namespace, cookies_path: Path, pdf_dir: Path) -> OrderStore:
    cookies = load_cookies(cookies_path)
    client = AliExpressMtopClient(
        cookies,
        ship_to_country=config.ship_to_country,
        lang=config.lang,
        time_zone=config.time_zone,
    )
    store = OrderStore()
    try:
        client.bootstrap()
        oldest_seen: Optional[date] = None
        last_reported_page_index: Optional[int] = None
        previous_list_payload: Optional[dict[str, Any]] = None
        for page_index in range(1, config.max_pages + 1):
            before_count = len(store.orders)
            if page_index == 1:
                payload = client.fetch_order_list_page(page_index, config.page_size)
            else:
                if previous_list_payload is None:
                    raise RuntimeError("Missing previous order list payload for pagination.")
                payload = client.fetch_order_list_page_more(
                    previous_list_payload,
                    page_index,
                    config.page_size,
                )
            previous_list_payload = payload
            metadata = parse_order_list_payload(payload, store)
            after_count = len(store.orders)
            new_orders = after_count - before_count
            page_orders = list(store.orders.values())
            parsed_dates = [resolve_order_date(order) for order in page_orders]
            parsed_dates = [item for item in parsed_dates if item is not None]
            if parsed_dates:
                oldest_seen = min(parsed_dates)
            reported_page_index = parse_optional_int(metadata.get("pageIndex"))
            print(
                f"Fetched order list page {page_index}: "
                f"{after_count} unique orders collected ({new_orders} new on this page)."
            )
            stop_reason = repeated_pagination_reason(
                page_index, new_orders, reported_page_index, last_reported_page_index
            )
            if stop_reason:
                print(f"Stopping pagination at page {page_index}: {stop_reason}")
                break
            if reported_page_index is not None:
                last_reported_page_index = reported_page_index
            if not metadata.get("hasMore"):
                break
            if oldest_seen and oldest_seen < parse_iso_date(config.start_date):
                break

        filtered_orders = store.filtered(
            parse_iso_date(config.start_date),
            parse_iso_date(config.end_date),
        )
        for bundle in filtered_orders:
            detail_payload = client.fetch_order_detail(bundle.order_id)
            parse_order_detail_payload(detail_payload, store)
            try:
                invoice_payload = client.fetch_invoice_info(bundle.order_id)
                parse_invoice_info_payload(invoice_payload, bundle.order_id, store)
            except Exception:
                continue

        if config.download_pdfs:
            for bundle in store.filtered(
                parse_iso_date(config.start_date),
                parse_iso_date(config.end_date),
            ):
                for index, invoice in enumerate(bundle.invoice_info_list):
                    pdf_path = invoice_pdf_path(pdf_dir, bundle.order_id, invoice, index)
                    if pdf_path.exists():
                        store.add_invoice_pdf_path(bundle.order_id, pdf_path)
                        continue
                    try:
                        pdf_content = client.fetch_invoice_file(bundle.order_id, invoice)
                    except Exception:
                        continue
                    if maybe_write_pdf(pdf_path, pdf_content):
                        store.add_invoice_pdf_path(bundle.order_id, pdf_path)

        save_cookies(cookies_path, client.export_cookies())
    finally:
        client.close()
    return store


def parse_order_id_from_request(entry: dict[str, Any]) -> str:
    url = entry.get("request", {}).get("url", "")
    query = parse_qs(urlparse(url).query)
    if "data" in query:
        try:
            payload = json.loads(query["data"][0])
            if "tradeOrderId" in payload:
                return str(payload["tradeOrderId"])
        except json.JSONDecodeError:
            pass
    post = entry.get("request", {}).get("postData", {})
    for param in post.get("params", []):
        if param.get("name") == "data":
            payload = json.loads(param.get("value", "{}"))
            return str(payload.get("orderId", ""))
    return ""


def export_from_hars(config: argparse.Namespace, pdf_dir: Path) -> OrderStore:
    store = OrderStore()
    for path_str in config.input_har:
        path = Path(path_str).expanduser().resolve()
        with path.open("r", encoding="utf-8") as handle:
            entries = json.load(handle).get("log", {}).get("entries", [])
        for entry in entries:
            url = entry.get("request", {}).get("url", "")
            content = entry.get("response", {}).get("content", {}) or {}
            text = content.get("text") or ""
            if not text:
                continue
            if "mtop.aliexpress.trade.buyer.order.list" in url:
                parse_order_list_payload(parse_jsonish_text(text), store)
            elif "mtop.aliexpress.trade.buyer.order.detail" in url:
                parse_order_detail_payload(parse_jsonish_text(text), store)
            elif "mtop.global.finance.taxation.invoice.getinvoiceinfolistbytype" in url:
                order_id = parse_order_id_from_request(entry)
                if order_id:
                    parse_invoice_info_payload(parse_jsonish_text(text), order_id, store)
            elif (
                config.download_pdfs
                and "mtop.global.finance.taxation.invoice.getinvoicefilecontent" in url
            ):
                order_id = parse_order_id_from_request(entry)
                payload = parse_jsonish_text(text)
                data = ensure_successful_mtop(payload, "invoice.file")
                content_b64 = data.get("data", {}).get("content", "")
                if not order_id or not content_b64:
                    continue
                post = entry.get("request", {}).get("postData", {})
                invoice_data = {}
                for param in post.get("params", []):
                    if param.get("name") == "data":
                        invoice_data = json.loads(param.get("value", "{}"))
                        break
                pdf_path = invoice_pdf_path(
                    pdf_dir,
                    order_id,
                    {
                        "invoiceId": invoice_data.get("invoiceId"),
                        "invoiceNo": invoice_data.get("invoiceNo"),
                    },
                    0,
                )
                if maybe_write_pdf(pdf_path, base64.b64decode(content_b64)):
                    store.add_invoice_pdf_path(order_id, pdf_path)
    return store


def run_export(config: argparse.Namespace) -> int:
    if not config.start_date or not config.end_date:
        raise SystemExit("--start-date and --end-date are required unless you use --setup.")
    start = parse_iso_date(config.start_date)
    end = parse_iso_date(config.end_date)
    if end < start:
        raise SystemExit("--end-date must be on or after --start-date.")

    output_dir = Path(config.output_dir).expanduser().resolve()
    csv_path = build_csv_path(output_dir, start, end)
    order_lines_csv_path = build_order_lines_csv_path(output_dir, start, end)
    xlsx_path = build_xlsx_path(output_dir, start, end)
    pdf_dir = build_pdf_dir(output_dir)
    cookies_path = Path(config.cookies_path).expanduser().resolve()

    imported_cookies: list[dict[str, Any]] = []
    if config.import_har:
        har_paths = [Path(path).expanduser().resolve() for path in config.import_har]
        imported_cookies.extend(import_cookies_from_hars(har_paths))
    if config.firefox_profile:
        imported_cookies.extend(
            import_cookies_from_firefox(Path(config.firefox_profile).expanduser().resolve())
        )
    if imported_cookies:
        save_cookies(cookies_path, imported_cookies)
        print(f"Stored AliExpress cookies in {cookies_path}")

    if config.input_har:
        store = export_from_hars(config, pdf_dir)
    else:
        if not cookies_path.exists():
            raise SystemExit(
                "No stored AliExpress cookies found.\n"
                "Import them first with either:\n"
                "  --import-har path/to/file.har\n"
                "  --firefox-profile path/to/FirefoxProfile\n"
            )
        store = export_live(config, cookies_path, pdf_dir)

    bundles = store.filtered(start, end)
    write_orders_csv(csv_path, bundles)
    write_order_lines_csv(order_lines_csv_path, bundles)
    if config.xlsx:
        write_xlsx(xlsx_path, bundles)

    print(f"CSV written to {csv_path}")
    print(f"Order-line CSV written to {order_lines_csv_path}")
    if config.xlsx:
        print(f"XLSX written to {xlsx_path}")
    if config.download_pdfs:
        print(f"PDF directory: {pdf_dir}")
    if cookies_path.exists():
        print(f"Stored auth cookies: {cookies_path}")
    print(f"Exported orders: {len(bundles)}")
    return 0


def main(argv: Optional[list[str]] = None) -> int:
    config = parse_args(argv)
    if config.setup:
        return run_interactive_setup(config)
    return run_export(config)


if __name__ == "__main__":
    raise SystemExit(main())
